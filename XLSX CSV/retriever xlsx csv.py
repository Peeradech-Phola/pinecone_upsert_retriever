import os
import dotenv
from pinecone import Pinecone, ServerlessSpec
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# โหลด environment variables จากไฟล์ .env
dotenv.load_dotenv()

# อ่านค่า environment variables
PINECONE_API_KEY = os.getenv("PINECONE_API_KEY")
PINECONE_INDEX_NAME = os.getenv("PINECONE_INDEX_NAME")
NAMESPACE = os.getenv("NAMESPACE")

# สร้างอินสแตนซ์ของ Pinecone
pc = Pinecone(api_key=PINECONE_API_KEY)

# ตรวจสอบว่าดัชนีมีอยู่แล้ว
if PINECONE_INDEX_NAME not in pc.list_indexes().names():
    raise ValueError(f"Index '{PINECONE_INDEX_NAME}' does not exist.")

# เข้าถึงดัชนี
index = pc.Index(PINECONE_INDEX_NAME)

# ดึงข้อมูลจาก Pinecone
def fetch_pinecone_data(namespace, top_k=20):
    dummy_vector = [0.0] * 1024  # ใช้ dummy vector
    response = index.query(vector=dummy_vector, top_k=top_k, namespace=namespace, include_metadata=True)
    return response

# ประมวลผลผลลัพธ์
response = fetch_pinecone_data(NAMESPACE)
data = []
for match in response["matches"]:
    metadata = match.get("metadata", {})
    if metadata.get("filename") == "dummy filename here":  ## Dummy file name here
        text = metadata.get("text", "")

        # แยกข้อความเป็นหมวดหมู่ (สมมุติว่าข้อความมีโครงสร้างแบบ key-value แยกด้วย colon)
        text_parts = text.split("\n")
        entry = {}
        for part in text_parts:
            if ":" in part:
                key, value = part.split(":", 1)
                entry[key.strip()] = value.strip()
        data.append(entry)

# แปลงข้อมูลเป็น DataFrame
df = pd.DataFrame(data)

# สร้างไฟล์ Excel ใหม่
output_file = r"Path Output/exaple.xlsx"  ## Path Output
workbook = Workbook()
sheet = workbook.active
sheet.title = "Pinecone Output"

# กำหนดหัวข้อ (Header)
headers = df.columns.tolist()
for col_num, header in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# เติมข้อมูลในไฟล์ Excel
for row_num, row in enumerate(df.itertuples(index=False), start=2):
    for col_num, value in enumerate(row, start=1):
        cell = sheet.cell(row=row_num, column=col_num)
        cell.value = value

# ปรับความกว้างของคอลัมน์
for col_num, column_cells in enumerate(sheet.columns, start=1):
    max_length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
    sheet.column_dimensions[sheet.cell(row=1, column=col_num).column_letter].width = max_length + 2

# บันทึกไฟล์
workbook.save(output_file)
print(f"Data saved to {output_file}")
# แปลงข้อมูลเป็น DataFrame
df = pd.DataFrame(data)

# แสดงข้อมูลทั้งหมดที่ดึงออกมาใน console
print("Fetched Data:")
print(df.to_string(index=False))